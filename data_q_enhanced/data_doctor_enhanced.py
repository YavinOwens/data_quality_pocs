import pandas as pd
import os
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from difflib import SequenceMatcher
from typing import Optional, List, Tuple, Dict
from collections import defaultdict
import glob
from IPython.display import display
from concurrent.futures import ThreadPoolExecutor, as_completed

class DataDoctor:
    """
    A class to handle data quality checks and template generation.
    """

    @staticmethod
    def clean_column_names(df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean column names by replacing non-alphanumeric characters with underscores and converting to lowercase.
        
        Args:
            df (pd.DataFrame): Dataframe to clean column names.
        
        Returns:
            pd.DataFrame: Dataframe with cleaned column names.
        """
        df.columns = [re.sub(r'\W+', '_', col).lower() for col in df.columns]
        return df

    @staticmethod
    def read_data_quality_template(excel_file_path: str) -> pd.DataFrame:
        """
        Read the data quality template from an Excel file.
        
        Args:
            excel_file_path (str): Path to the Excel file.
        
        Returns:
            pd.DataFrame: DataFrame containing the data quality template.
        """
        df_template = pd.read_excel(excel_file_path, sheet_name='Data Quality Checks', skiprows=1)
        return df_template

    @staticmethod
    def assess_completeness(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
        """
        Assess completeness for a specific column in the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe containing the data.
            column_name (str): Name of the column to assess.
        
        Returns:
            pd.DataFrame: DataFrame containing completeness assessment results.
        """
        total_rows = len(df)
        missing_values = df[column_name].isnull().sum()
        non_missing_values = total_rows - missing_values
        completeness_percentage = (non_missing_values / total_rows) * 100

        completeness_df = pd.DataFrame({
            'Column Name': [column_name],
            'Total Rows': [total_rows],
            'Missing Values': [missing_values],
            'Non-Missing Values': [non_missing_values],
            'Completeness (%)': [round(completeness_percentage, 2)]
        })

        return completeness_df

    @staticmethod
    def assess_uniqueness(df: pd.DataFrame, column_name: str) -> pd.DataFrame:
        """
        Assess uniqueness for a specific column in the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe containing the data.
            column_name (str): Name of the column to assess.
        
        Returns:
            pd.DataFrame: DataFrame containing uniqueness assessment results.
        """
        total_rows = len(df)
        unique_values = df[column_name].nunique()
        duplicate_values = total_rows - unique_values
        uniqueness_percentage = (unique_values / total_rows) * 100

        uniqueness_df = pd.DataFrame({
            'Column Name': [column_name],
            'Total Rows': [total_rows],
            'Unique Values': [unique_values],
            'Duplicate Values': [duplicate_values],
            'Uniqueness (%)': [round(uniqueness_percentage, 2)]
        })

        return uniqueness_df

    @staticmethod
    def assess_consistency(df: pd.DataFrame, column_name: str, pattern: Optional[str] = None, valid_values: Optional[List[str]] = None) -> pd.DataFrame:
        """
        Assess consistency for a specific column in the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe containing the data.
            column_name (str): Name of the column to assess.
            pattern (str, optional): Regex pattern to validate the column values.
            valid_values (List[str], optional): List of valid values to check against.
        
        Returns:
            pd.DataFrame: DataFrame containing consistency assessment results.
        """
        total_rows = len(df)
        consistent_values = 0

        if pattern:
            try:
                re.compile(pattern)
                consistent_values = df[column_name].apply(lambda x: bool(re.match(pattern, str(x)))).sum()
            except re.error as e:
                print(f"Error in regex pattern '{pattern}' for column '{column_name}': {e}")
        elif valid_values:
            consistent_values = df[column_name].isin(valid_values).sum()
        else:
            consistent_values = total_rows  # Assuming all values are consistent if no pattern or valid values provided.

        inconsistent_values = total_rows - consistent_values
        consistency_percentage = (consistent_values / total_rows) * 100

        consistency_df = pd.DataFrame({
            'Column Name': [column_name],
            'Total Rows': [total_rows],
            'Consistent Values': [consistent_values],
            'Inconsistent Values': [inconsistent_values],
            'Consistency (%)': [round(consistency_percentage, 2)]
        })

        return consistency_df

    @staticmethod
    def assess_validity(df: pd.DataFrame, column_name: str, valid_range: Optional[Tuple[str, str]] = None) -> pd.DataFrame:
        """
        Assess validity for a specific column in the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe containing the data.
            column_name (str): Name of the column to assess.
            valid_range (Tuple[str, str], optional): Tuple specifying the valid range of values (start, end).
        
        Returns:
            pd.DataFrame: DataFrame containing validity assessment results.
        """
        total_rows = len(df)
        if valid_range:
            try:
                min_val, max_val = valid_range
                valid_values = df[column_name].apply(lambda x: pd.to_datetime(min_val) <= pd.to_datetime(x) <= pd.to_datetime(max_val) if pd.notnull(x) else False).sum()
            except Exception as e:
                print(f"Error in date range '{valid_range}' for column '{column_name}': {e}")
                valid_values = 0
        else:
            valid_values = total_rows  # Assuming all values are valid if no range provided.

        invalid_values = total_rows - valid_values
        validity_percentage = (valid_values / total_rows) * 100

        validity_df = pd.DataFrame({
            'Column Name': [column_name],
            'Total Rows': [total_rows],
            'Valid Values': [valid_values],
            'Invalid Values': [invalid_values],
            'Validity (%)': [round(validity_percentage, 2)]
        })

        return validity_df

    @staticmethod
    def assess_accuracy(df: pd.DataFrame, column_name: str, reference_values: List[str]) -> pd.DataFrame:
        """
        Assess accuracy for a specific column in the dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe containing the data.
            column_name (str): Name of the column to assess.
            reference_values (List[str]): List of reference values to check against.
        
        Returns:
            pd.DataFrame: DataFrame containing accuracy assessment results.
        """
        total_rows = len(df)
        accurate_values = df[column_name].isin(reference_values).sum()
        inaccurate_values = total_rows - accurate_values
        accuracy_percentage = (accurate_values / total_rows) * 100

        accuracy_df = pd.DataFrame({
            'Column Name': [column_name],
            'Total Rows': [total_rows],
            'Accurate Values': [accurate_values],
            'Inaccurate Values': [inaccurate_values],
            'Accuracy (%)': [round(accuracy_percentage, 2)]
        })

        return accuracy_df

    @staticmethod
    def similar(a: str, b: str) -> float:
        """
        Calculate similarity ratio between two strings.
        
        Args:
            a (str): First string.
            b (str): Second string.
        
        Returns:
            float: Similarity ratio.
        """
        return SequenceMatcher(None, a, b).ratio()

    @staticmethod
    def is_pii(column_name: str) -> bool:
        """
        Check if a column name indicates personally identifiable information (PII).
        
        Args:
            column_name (str): Column name to check.
        
        Returns:
            bool: True if column name indicates PII, False otherwise.
        """
        pii_keywords = ["name", "dob", "date of birth", "age", "contact number"]
        for keyword in pii_keywords:
            if DataDoctor.similar(column_name.lower(), keyword) > 0.8:
                return True
        return False

    @staticmethod
    def read_all_structured_files(directory_path: str) -> List[Tuple[str, pd.DataFrame]]:
        """
        Read all CSV and Excel files from a directory and return their data as dataframes.
        
        Args:
            directory_path (str): Path to the directory containing files.
        
        Returns:
            List[Tuple[str, pd.DataFrame]]: List of tuples containing file paths and dataframes.
        """
        all_files = glob.glob(os.path.join(directory_path, "*.csv")) + glob.glob(os.path.join(directory_path, "*.xlsx"))
        all_sheets = []
        
        for file_path in all_files:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path)
                df = DataDoctor.clean_column_names(df)
                all_sheets.append((file_path, df))
            elif file_path.endswith('.xlsx'):
                xls = pd.ExcelFile(file_path)
                for sheet_name in xls.sheet_names:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    df = DataDoctor.clean_column_names(df)
                    all_sheets.append((f"{file_path} - {sheet_name}", df))
        
        return all_sheets

    @staticmethod
    def find_critical_elements(all_sheets: List[Tuple[str, pd.DataFrame]]) -> Dict[str, List[str]]:
        """
        Find critical elements (columns) that appear in multiple files.
        
        Args:
            all_sheets (List[Tuple[str, pd.DataFrame]]): List of tuples containing file paths and dataframes.
        
        Returns:
            Dict[str, List[str]]: Dictionary with column names as keys and list of file paths as values.
        """
        column_files_map = defaultdict(list)
        for file_path, df in all_sheets:
            for column in df.columns:
                column_files_map[column].append(file_path)
        
        critical_elements = {column: files for column, files in column_files_map.items() if len(files) > 1}
        return critical_elements

    def configure_quality_check(self, csv_file_path: str, excel_file_path: Optional[str] = None) -> None:
        """
        Configure quality check and create an Excel template if it doesn't already exist.
        
        Args:
            csv_file_path (str): Path to the CSV file for which to configure the quality check.
            excel_file_path (Optional[str]): Path to save the Excel template. If None, saves in the current directory.
        """
        if not excel_file_path:
            excel_file_path = os.path.join(os.getcwd(), 'data_quality_checks_template.xlsx')

        if os.path.exists(excel_file_path):
            print(f"File '{excel_file_path}' already exists. Not overwriting.")
            return

        all_sheets = self.read_all_structured_files(os.path.dirname(csv_file_path))
        critical_elements = self.find_critical_elements(all_sheets)
        
        df = pd.read_csv(csv_file_path)
        df = self.clean_column_names(df)
        column_names = df.columns.tolist()

        pii_flags = []
        critical_data_elements = []
        
        for column in column_names:
            if self.is_pii(column):
                similar_columns = [col for sheet in all_sheets for col in sheet[1].columns if self.similar(col.lower(), column.lower()) > 0.8]
                description = ', '.join(set(similar_columns))
                pii_flags.append(f"Yes, description: {description}")
            else:
                pii_flags.append("No")
            
            if column in critical_elements:
                critical_data_elements.append(f"Yes, files: {', '.join(critical_elements[column])}")
            else:
                critical_data_elements.append("No")
        
        data_quality_checks_df = pd.DataFrame({
            "column_names": column_names,
            "PII_Flag": pii_flags,
            "test_completeness": ["Not Assessed" for _ in column_names],  # Set default value to "Not Assessed"
            "test_uniqueness": ["" for _ in column_names],
            "test_timeliness": ["" for _ in column_names],
            "test_consistency": ["" for _ in column_names],
            "test_accuracy": ["" for _ in column_names],
            "test_validity": ["" for _ in column_names],
            "critical_data_element": critical_data_elements,
            "pattern": ["" for _ in column_names],  # Add columns for pattern and valid values
            "valid_values": ["" for _ in column_names],
            "valid_range": ["" for _ in column_names],  # Add column for valid range
            "reference_values": ["" for _ in column_names]  # Add column for reference values
        })

        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            data_quality_checks_df.to_excel(writer, sheet_name='Data Quality Checks', index=False, startrow=1)

        # Load the workbook to add the description
        workbook = load_workbook(excel_file_path)
        sheet = workbook['Data Quality Checks']
        
        # Add description at the top
        description = (f"File Name: {os.path.basename(csv_file_path)}\n"
                       "Please provide 'Yes' or 'No' in the columns below for each data quality check.")
        sheet['A1'] = description
        sheet.merge_cells('A1:H1')
        sheet['A1'].alignment = Alignment(wrap_text=True, vertical='center')

        # Adjust column widths
        for col in range(1, sheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    print(f"Error adjusting width for column {column}: {e}")
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        workbook.save(excel_file_path)
        print(f"Excel file '{excel_file_path}' created successfully with instructions.")

    def evaluate_column_quality(self, df_data: pd.DataFrame, df_template: pd.DataFrame, column_name: str, failing_records: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        """
        Evaluate data quality for a single column.
        
        Args:
            df_data (pd.DataFrame): Dataframe containing the data.
            df_template (pd.DataFrame): Dataframe containing the data quality template.
            column_name (str): Name of the column to assess.
            failing_records (pd.DataFrame): Dataframe to collect failing records.
        
        Returns:
            Dict[str, pd.DataFrame]: Dictionary containing assessment results for completeness, uniqueness, consistency, validity, and accuracy.
        """
        results = {
            'completeness': pd.DataFrame(),
            'uniqueness': pd.DataFrame(),
            'consistency': pd.DataFrame(),
            'validity': pd.DataFrame(),
            'accuracy': pd.DataFrame()
        }
        
        row = df_template[df_template['column_names'] == column_name].iloc[0]
        test_completeness = str(row['test_completeness']).strip().lower() if pd.notna(row['test_completeness']) else 'not assessed'
        test_uniqueness = str(row['test_uniqueness']).strip().lower() if pd.notna(row['test_uniqueness']) else 'not assessed'
        test_consistency = str(row['test_consistency']).strip().lower() if pd.notna(row['test_consistency']) else 'not assessed'
        test_validity = str(row['test_validity']).strip().lower() if pd.notna(row['test_validity']) else 'not assessed'
        test_accuracy = str(row['test_accuracy']).strip().lower() if pd.notna(row['test_accuracy']) else 'not assessed'
        
        if test_completeness == 'yes':
            completeness_df = self.assess_completeness(df_data, column_name)
            if not completeness_df.empty:
                results['completeness'] = completeness_df
            missing_records = df_data[df_data[column_name].isnull()]
            if not missing_records.empty:
                missing_records = missing_records.copy()
                if 'Failed Test' not in missing_records:
                    missing_records['Failed Test'] = ''
                if 'Reason' not in missing_records:
                    missing_records['Reason'] = ''
                missing_records.loc[:, 'Failed Test'] += ';Completeness'
                missing_records.loc[:, 'Reason'] += ';Missing Values'
                failing_records = pd.concat([failing_records, missing_records])
        
        if test_uniqueness == 'yes':
            uniqueness_df = self.assess_uniqueness(df_data, column_name)
            if not uniqueness_df.empty:
                results['uniqueness'] = uniqueness_df
            duplicate_records = df_data[df_data.duplicated(column_name, keep=False)]
            if not duplicate_records.empty:
                duplicate_records = duplicate_records.copy()
                if 'Failed Test' not in duplicate_records:
                    duplicate_records['Failed Test'] = ''
                if 'Reason' not in duplicate_records:
                    duplicate_records['Reason'] = ''
                duplicate_records.loc[:, 'Failed Test'] += ';Uniqueness'
                duplicate_records.loc[:, 'Reason'] += ';Duplicate Values'
                failing_records = pd.concat([failing_records, duplicate_records])
        
        if test_consistency == 'yes':
            pattern = row['pattern'] if 'pattern' in row and pd.notna(row['pattern']) else None
            valid_values = row['valid_values'].split(';') if 'valid_values' in row and pd.notna(row['valid_values']) else None
            consistency_df = self.assess_consistency(df_data, column_name, pattern, valid_values)
            if not consistency_df.empty:
                results['consistency'] = consistency_df
            if pattern:
                try:
                    re.compile(pattern)
                    inconsistent_records = df_data[~df_data[column_name].apply(lambda x: bool(re.match(pattern, str(x))))]
                except re.error as e:
                    print(f"Error in regex pattern '{pattern}' for column '{column_name}': {e}")
                    inconsistent_records = pd.DataFrame()
            elif valid_values:
                inconsistent_records = df_data[~df_data[column_name].isin(valid_values)]
            if not inconsistent_records.empty:
                inconsistent_records = inconsistent_records.copy()
                if 'Failed Test' not in inconsistent_records:
                    inconsistent_records['Failed Test'] = ''
                if 'Reason' not in inconsistent_records:
                    inconsistent_records['Reason'] = ''
                inconsistent_records.loc[:, 'Failed Test'] += ';Consistency'
                inconsistent_records.loc[:, 'Reason'] += ';Invalid Pattern or Value'
                failing_records = pd.concat([failing_records, inconsistent_records])
        
        if test_validity == 'yes':
            valid_range = tuple(row['valid_range'].split(';')) if 'valid_range' in row and pd.notna(row['valid_range']) else None
            validity_df = self.assess_validity(df_data, column_name, valid_range)
            if not validity_df.empty:
                results['validity'] = validity_df
            if valid_range:
                min_val, max_val = valid_range
                invalid_records = df_data[~df_data[column_name].apply(lambda x: pd.to_datetime(min_val) <= pd.to_datetime(x) <= pd.to_datetime(max_val) if pd.notnull(x) else False)]
            if not invalid_records.empty:
                invalid_records = invalid_records.copy()
                if 'Failed Test' not in invalid_records:
                    invalid_records['Failed Test'] = ''
                if 'Reason' not in invalid_records:
                    invalid_records['Reason'] = ''
                invalid_records.loc[:, 'Failed Test'] += ';Validity'
                invalid_records.loc[:, 'Reason'] += ';Out of Range'
                failing_records = pd.concat([failing_records, invalid_records])
        
        if test_accuracy == 'yes':
            reference_values = row['reference_values'].split(';') if 'reference_values' in row and pd.notna(row['reference_values']) else None
            accuracy_df = self.assess_accuracy(df_data, column_name, reference_values)
            if not accuracy_df.empty:
                results['accuracy'] = accuracy_df
            inaccurate_records = df_data[~df_data[column_name].isin(reference_values)]
            if not inaccurate_records.empty:
                inaccurate_records = inaccurate_records.copy()
                if 'Failed Test' not in inaccurate_records:
                    inaccurate_records['Failed Test'] = ''
                if 'Reason' not in inaccurate_records:
                    inaccurate_records['Reason'] = ''
                inaccurate_records.loc[:, 'Failed Test'] += ';Accuracy'
                inaccurate_records.loc[:, 'Reason'] += ';Invalid Reference Value'
                failing_records = pd.concat([failing_records, inaccurate_records])

        return results

    def evaluate_data_quality(self, data_file_path: str, template_file_path: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Evaluate data quality based on a template.
        
        Args:
            data_file_path (str): Path to the data file (.csv or .xlsx).
            template_file_path (str): Path to the template file (.xlsx).
        
        Returns:
            Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]: DataFrames containing completeness, uniqueness, consistency, validity, accuracy assessment results, and failing records.
        """
        df_template = self.read_data_quality_template(template_file_path)

        if data_file_path.endswith('.csv'):
            df_data = pd.read_csv(data_file_path)
        elif data_file_path.endswith('.xlsx'):
            df_data = pd.read_excel(data_file_path)
        else:
            raise ValueError("Unsupported file format. Please use .csv or .xlsx files.")

        completeness_results = pd.DataFrame(columns=['Column Name', 'Total Rows', 'Missing Values', 'Non-Missing Values', 'Completeness (%)'])
        uniqueness_results = pd.DataFrame(columns=['Column Name', 'Total Rows', 'Unique Values', 'Duplicate Values', 'Uniqueness (%)'])
        consistency_results = pd.DataFrame(columns=['Column Name', 'Total Rows', 'Consistent Values', 'Inconsistent Values', 'Consistency (%)'])
        validity_results = pd.DataFrame(columns=['Column Name', 'Total Rows', 'Valid Values', 'Invalid Values', 'Validity (%)'])
        accuracy_results = pd.DataFrame(columns=['Column Name', 'Total Rows', 'Accurate Values', 'Inaccurate Values', 'Accuracy (%)'])

        failing_records = pd.DataFrame(columns=df_data.columns.tolist() + ['Failed Test', 'Reason'])

        if not df_data.empty:
            df_data = self.clean_column_names(df_data)
            
            with ThreadPoolExecutor() as executor:
                future_to_column = {executor.submit(self.evaluate_column_quality, df_data, df_template, column, failing_records): column for column in df_template['column_names']}
                
                for future in as_completed(future_to_column):
                    column_name = future_to_column[future]
                    try:
                        result = future.result()
                        if not result['completeness'].empty:
                            completeness_results = pd.concat([completeness_results, result['completeness']], ignore_index=True)
                        if not result['uniqueness'].empty:
                            uniqueness_results = pd.concat([uniqueness_results, result['uniqueness']], ignore_index=True)
                        if not result['consistency'].empty:
                            consistency_results = pd.concat([consistency_results, result['consistency']], ignore_index=True)
                        if not result['validity'].empty:
                            validity_results = pd.concat([validity_results, result['validity']], ignore_index=True)
                        if not result['accuracy'].empty:
                            accuracy_results = pd.concat([accuracy_results, result['accuracy']], ignore_index=True)
                    except Exception as e:
                        print(f"Error processing column '{column_name}': {e}")

        else:
            print("Warning: The data file is empty.")

        failing_records['Failed Test'] = failing_records['Failed Test'].str.strip(';')
        failing_records['Reason'] = failing_records['Reason'].str.strip(';')

        return completeness_results, uniqueness_results, consistency_results, validity_results, accuracy_results, failing_records

    def display_failing_records(self, failing_records: pd.DataFrame) -> None:
        """
        Display failing records in a dataframe format.
        
        Args:
            failing_records (pd.DataFrame): DataFrame containing the failing records.
        """
        if failing_records.empty:
            print("No failing records to display.")
        else:
            display(failing_records)

    def generate_data_dictionary(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate a data dictionary for the given dataframe.
        
        Args:
            df (pd.DataFrame): Dataframe to generate data dictionary for.
        
        Returns:
            pd.DataFrame: Data dictionary with standard and enhanced labels.
        """
        data_types = []
        pii_flags = []
        critical_elements = []
        data_type_presentation = []

        for column in df.columns:
            if pd.api.types.is_numeric_dtype(df[column]):
                data_types.append('Number')
                data_type_presentation.append('N')
            elif pd.api.types.is_string_dtype(df[column]):
                data_types.append('String')
                data_type_presentation.append('A')
            elif pd.api.types.is_datetime64_any_dtype(df[column]):
                data_types.append('Datetime')
                data_type_presentation.append('DT')
            else:
                data_types.append('Unknown')
                data_type_presentation.append('AN')

            if self.is_pii(column):
                pii_flags.append("Yes")
            else:
                pii_flags.append("No")

            # Assuming critical elements were identified earlier
            critical_elements.append("No")  # Placeholder, needs actual check from earlier logic

        data_dictionary_df = pd.DataFrame({
            'Column Name': df.columns,
            'Data Type': data_types,
            'PII': pii_flags,
            'Critical Element': critical_elements,
            'Data Type Presentation': data_type_presentation
        })

        return data_dictionary_df

# Example usage:
if __name__ == "__main__":
    data_file_path = 'your_data_file.csv'  # Update with the actual path
    template_file_path = 'data_quality_checks_template.xlsx'

    # Create an instance of DataDoctor
    data_doctor = DataDoctor()

    # Check if the template file exists, and create if it doesn't
    data_doctor.configure_quality_check(data_file_path, template_file_path)

    # Evaluate data quality based on the template
    completeness_results, uniqueness_results, consistency_results, validity_results, accuracy_results, failing_records = data_doctor.evaluate_data_quality(data_file_path, template_file_path)

    # Display the completeness results DataFrame
    print("Completeness Results DataFrame:")
    display(completeness_results)

    # Display the uniqueness results DataFrame
    print("Uniqueness Results DataFrame:")
    display(uniqueness_results)

    # Display the consistency results DataFrame
    print("Consistency Results DataFrame:")
    display(consistency_results)

    # Display the validity results DataFrame
    print("Validity Results DataFrame:")
    display(validity_results)

    # Display the accuracy results DataFrame
    print("Accuracy Results DataFrame:")
    display(accuracy_results)

    # Display the failing records DataFrame
    data_doctor.display_failing_records(failing_records)

    # Generate and display the data dictionary
    print("Data Dictionary:")
    df_data = pd.read_csv(data_file_path)
    data_dictionary = data_doctor.generate_data_dictionary(df_data)
    display(data_dictionary)
